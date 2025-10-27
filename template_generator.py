from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from datetime import datetime, date
from pathlib import Path
from typing import Dict, List, Union
from openpyxl import load_workbook

NumberLike = Union[float, int, str]

def _clean_to_float(v: NumberLike) -> float:
    if isinstance(v, (int, float)): return float(v)
    if isinstance(v, str):
        try: return float(v.replace("$","").replace(",","").strip())
        except: return 0.0
    return 0.0

def _coerce_year_key(k) -> int | None:
    try: return int(str(k).strip())
    except: return None

def _elapsed_remaining_le(le_months: int, le_report_date: str) -> tuple[int,int,int]:
    # (elapsed_months, remaining_le_months, remaining_le_years)
    le_dt = datetime.strptime(le_report_date, "%Y-%m-%d")
    today = date.today()
    elapsed = (today.year - le_dt.year) * 12 + (today.month - le_dt.month)
    rem = max(int(le_months) - elapsed, 0)
    rem_years = (rem + 11) // 12
    return elapsed, rem, rem_years

def _age_today(dob: str, le_report_date: str, elapsed_months: int) -> int:
    dob_dt = datetime.strptime(dob, "%Y-%m-%d")
    le_dt = datetime.strptime(le_report_date, "%Y-%m-%d")
    return int((le_dt - dob_dt).days / 365.25 + elapsed_months / 12)

def _normalize_month_map(monthly_premiums: Dict) -> Dict[int, List[float]]:
    # {year: [ ... months ... ]}, keep only int year keys; coerce values to float list
    out: Dict[int, List[float]] = {}
    for k, v in (monthly_premiums or {}).items():
        y = _coerce_year_key(k)
        if y is None: continue
        if not isinstance(v, (list, tuple)):
            out[y] = []
            continue
        out[y] = [_clean_to_float(x) for x in v]
    return out

def _month_value_for(year: int, month_idx0: int, year_map: Dict[int, List[float]]) -> float:
    """
    0-based month; align arrays to end of year:
    - if len(list) == 12: direct index
    - if len(list) < 12: last element = December; missing leading months are zeros
    """
    arr = year_map.get(year, [])
    L = len(arr)
    if L == 12:
        return arr[month_idx0]
    if L == 0:
        return 0.0
    start_at = 12 - L  # first filled month index
    if month_idx0 < start_at:
        return 0.0
    return arr[month_idx0 - start_at]

def _sum_next_n_months(start_year: int, start_month_idx0: int, n: int, year_map: Dict[int, List[float]]) -> float:
    total = 0.0
    y, m = start_year, start_month_idx0
    for _ in range(max(n, 0)):
        total += _month_value_for(y, m, year_map)
        m += 1
        if m == 12:
            m = 0
            y += 1
    return total

def _premiums_to_le(start_year: int, start_month_idx0: int, rem_le_months: int, year_map: Dict[int, List[float]]) -> float:
    if rem_le_months <= 0:
        return 0.0
    # include current month + next (rem_le_months-1)
    return _sum_next_n_months(start_year, start_month_idx0, rem_le_months, year_map)
# === end helpers ===


def _clean_to_float(v: NumberLike) -> float:
    """Convert numbers or currency-like strings to float; invalid -> 0.0"""
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        try:
            return float(v.replace("$", "").replace(",", "").strip())
        except Exception:
            return 0.0
    return 0.0

def _coerce_year_key(k) -> Union[int, None]:
    """Coerce year keys like '2025', '2025.0', 2025.0 -> 2025; return None if impossible."""
    if isinstance(k, int):
        return k
    if isinstance(k, float):
        return int(round(k))
    if isinstance(k, str):
        s = k.strip()
        try:
            return int(s)
        except Exception:
            try:
                return int(float(s))
            except Exception:
                return None
    return None

def _get_monthly_premium_frontfill_tail(
    mp_dict: Dict[int, List[NumberLike]],
    year: int,
    month_1based: int
) -> float:
    """
    Return the premium for (year, month). Each year maps to a list of monthly values.
    If a year has < 12 entries, assume the missing months are at the FRONT (Jan..),
    i.e., the provided values align to the END of the year (… Sep, Oct, Nov, Dec).
    Example: 9 entries -> treat Jan/Feb/Mar as 0, entries correspond to Apr..Dec.
    """
    lst = mp_dict.get(year)
    if not isinstance(lst, (list, tuple)):
        return 0.0

    clean = [_clean_to_float(x) for x in lst]
    n = len(clean)

    # Clamp month to [1..12]
    m = max(1, min(12, month_1based))

    if n >= 12:
        return clean[m - 1]

    # Front-fill zeros: provided values occupy the LAST n months of the year
    offset = 12 - n  # number of missing months at the start (treated as zeros)
    idx = (m - 1) - offset
    if idx < 0 or idx >= n:
        return 0.0
    return clean[idx]

def generate_return_template(
    insured_name: str,
    dob: str,
    carrier: str,
    le_months: int,
    le_report_date: str,
    death_benefit: float,
    investment: float,
    monthly_premiums: Dict[int, List[float]],
    output_filename: str
) -> str:
    # Parse dates / anchors
    dob_dt = datetime.strptime(dob, "%Y-%m-%d")
    le_report_dt = datetime.strptime(le_report_date, "%Y-%m-%d")
    today = date.today()

    # Elapsed/remaining LE
    elapsed_months = (today.year - le_report_dt.year) * 12 + today.month - le_report_dt.month
    remaining_le_months = max(le_months - elapsed_months, 0)
    remaining_le_years = (remaining_le_months + 11) // 12
    total_years = remaining_le_years + 3
    start_year = today.year

    # Approximate age at "today" anchored from LE report date + elapsed months
    age = int((le_report_dt - dob_dt).days / 365.25 + elapsed_months / 12)

    # Coerce year keys robustly and keep only valid ones
    monthly_premiums = {
        yk: v for k, v in (monthly_premiums or {}).items()
        if (yk := _coerce_year_key(k)) is not None
    }

    # Annual premium totals (robust to string inputs)
    annual_premiums: Dict[int, float] = {}
    for year, months in monthly_premiums.items():
        if not isinstance(months, (list, tuple)):
            annual_premiums[year] = 0.0
            continue
        annual_premiums[year] = sum(_clean_to_float(x) for x in months)

    # Load template and clear old output rows (keep headers up to row 6)
    wb = load_workbook("return_template_output.xlsx")
    ws = wb.active

    for _ in range(7, ws.max_row + 1):
        ws.delete_rows(7)

    # Header cells
    ws["B1"] = insured_name
    ws["B2"] = f"AGE: {age}"
    ws["B3"] = f"CARRIER: {carrier}"
    ws["E2"] = f"{remaining_le_months} MONTHS"
    ws["E3"] = death_benefit
    ws["E4"] = investment

    # === Auto-calc: next 3 months of premiums (including this month) -> E5 ===
    now_dt = datetime.now()
    cur_y, cur_m = now_dt.year, now_dt.month  # 1..12

    # Build (year, month) positions for current month + next two months, handling year wrap
    positions = []
    for off in range(3):
        y = cur_y + ((cur_m - 1 + off) // 12)
        m = ((cur_m - 1 + off) % 12) + 1
        positions.append((y, m))

    next_three_sum = sum(
        _get_monthly_premium_frontfill_tail(monthly_premiums, y, m) for (y, m) in positions
    )
    ws["E5"] = next_three_sum
    ws["E5"].number_format = '"$"#,##0.00'
    # === end auto-calc E5 ===

    # Year-by-year table
    cumulative = 0.0
    for i in range(total_years):
        year = start_year + i
        premium = float(annual_premiums.get(year, 0.0))
        cumulative += premium
        total_cost = float(investment) + cumulative
        profit = float(death_benefit) - total_cost
        simple_return = (profit / total_cost) if total_cost else 0.0
        acc_return = (simple_return / (i + 1)) if (i + 1) else 0.0

        marker = ""
        if i == remaining_le_years - 1:
            marker = "LE"
        elif i == remaining_le_years:
            marker = "LE+1"
        elif i == remaining_le_years + 1:
            marker = "LE+2"
        elif i == remaining_le_years + 2:
            marker = "LE+3"

        row = [
            year,
            premium,
            cumulative,
            total_cost,
            profit,
            simple_return,
            acc_return,
            marker
        ]
        ws.append(row)

        # Highlight LE row (bold + light blue fill)
        if i == remaining_le_years - 1:
            for col in range(2, 9):  # columns B..H
                cell = ws.cell(row=6 + i + 1, column=col)  # == row 7 + i
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

    # Number formats for the appended rows
    for row in range(7, 7 + total_years):
        for col in range(2, 6):  # B..E currency
            ws.cell(row=row, column=col).number_format = '"$"#,##0.00'
        for col in range(6, 8):  # F..G percentages
            ws.cell(row=row, column=col).number_format = '0.00%'

    # Copy the style from B6 into column A for all output rows
    ref_style = ws["B6"]._style
    for row in range(7, 7 + total_years):
        ws.cell(row=row, column=1)._style = ref_style

    # Clear any "LE Marker" header label in H6 to keep header area clean
    if ws["H6"].value == "LE Marker":
        ws["H6"].value = ""

    wb.save(output_filename)
    return output_filename

# === Resale generator (paste near the end of file) ===
def generate_resale_template(
    insured_name: str,
    dob: str,
    carrier: str,
    le_months: int,
    le_report_date: str,
    death_benefit: NumberLike,
    investment: NumberLike,              # client’s purchase price (matches E4 in purchase template)
    monthly_premiums: Dict,              # {year: [months]}
    output_filename: str = None
) -> str:
    today = date.today()
    this_year, this_month_idx0 = today.year, today.month - 1  # 0-based month
    elapsed, remaining_le_months, _ = _elapsed_remaining_le(int(le_months), le_report_date)
    age = _age_today(dob, le_report_date, elapsed)

    premium_map = _normalize_month_map(monthly_premiums)
    DB = _clean_to_float(death_benefit)
    COST0 = _clean_to_float(investment)

    # Load template from repo root (same folder as this .py or project root)
    here = Path(__file__).resolve().parent
    candidate_paths = [
        here / "Resale Template Sample.xlsx",
        Path.cwd() / "Resale Template Sample.xlsx",
    ]
    xlsx_path = next((p for p in candidate_paths if p.exists()), None)
    if xlsx_path is None:
        raise FileNotFoundError("Resale Template Sample.xlsx not found in repo. Place it next to template_generator.py or at project root.")

    wb = load_workbook(xlsx_path.as_posix())
    ws = wb.active

    # Headers
    ws["B1"].value = insured_name
    ws["B2"].value = f"AGE: {age}"
    ws["B3"].value = f"CARRIER: {carrier}"
    ws["F1"].value = today.strftime("%Y-%m-%d")
    ws["F2"].value = f"{remaining_le_months} MONTHS"
    ws["F3"].value = DB
    ws["F4"].value = COST0

    # Effective months (B8..B11)
    B8 = max(remaining_le_months - 24, 0)
    B9 = max(B8 - 12, 0)
    B10 = max(B9 - 12, 0)
    B11 = max(B10 - 12, 0)
    ws["B8"].value, ws["B9"].value, ws["B10"].value, ws["B11"].value = B8, B9, B10, B11

    # Cumulative premiums from today (C8..C11): 24/36/48/60 months
    C8 = _sum_next_n_months(this_year, this_month_idx0, 24, premium_map)
    C9 = _sum_next_n_months(this_year, this_month_idx0, 36, premium_map)
    C10 = _sum_next_n_months(this_year, this_month_idx0, 48, premium_map)
    C11 = _sum_next_n_months(this_year, this_month_idx0, 60, premium_map)
    ws["C8"].value, ws["C9"].value, ws["C10"].value, ws["C11"].value = C8, C9, C10, C11

    # Premiums to LE (baseline) and D8..D11 deltas
    P_LE = _premiums_to_le(this_year, this_month_idx0, remaining_le_months, premium_map)
    D8 = max(P_LE - C8, 0.0)
    D9 = max(P_LE - C9, 0.0)
    D10 = max(P_LE - C10, 0.0)
    D11 = max(P_LE - C11, 0.0)
    ws["D8"].value, ws["D9"].value, ws["D10"].value, ws["D11"].value = D8, D9, D10, D11

    # Resale price (Client 2 buys; 18% avg return to LE)
    def _resale_price(Bm, Dm):
        denom = (0.18 * max(Bm, 0)) + 12.0
        return ((12.0 * DB) - (12.0 * Dm)) / denom if denom > 0 else 0.0

    E8 = _resale_price(B8, D8)
    E9 = _resale_price(B9, D9)
    E10 = _resale_price(B10, D10)
    E11 = _resale_price(B11, D11)
    ws["E8"].value, ws["E9"].value, ws["E10"].value, ws["E11"].value = E8, E9, E10, E11

    # Client 1 proceeds (F) and annualized returns (G)
    F8, F9, F10, F11 = E8 - C8 - COST0, E9 - C9 - COST0, E10 - C10 - COST0, E11 - C11 - COST0
    ws["F8"].value, ws["F9"].value, ws["F10"].value, ws["F11"].value = F8, F9, F10, F11

    def _ann_return(Fm, Cm, yrs):
        base = COST0 + Cm
        return (Fm / base / yrs) if base > 0 and yrs > 0 else 0.0

    G8 = _ann_return(F8, C8, 2)
    G9 = _ann_return(F9, C9, 3)
    G10 = _ann_return(F10, C10, 4)
    G11 = _ann_return(F11, C11, 5)
    ws["G8"].value, ws["G9"].value, ws["G10"].value, ws["G11"].value = G8, G9, G10, G11

    # Formats
    for addr in ("F3","F4","C8","C9","C10","C11","D8","D9","D10","D11","E8","E9","E10","E11","F8","F9","F10","F11"):
        ws[addr].number_format = '"$"#,##0.00'
    for addr in ("G8","G9","G10","G11"):
        ws[addr].number_format = '0.00%'

    # Optional visual cue on B8:B11
    fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
    for r in (8,9,10,11):
        ws[f"B{r}"].fill = fill

    safe = insured_name.lower().replace(" ", "_")
    out = output_filename or f"resale_template_{safe}.xlsx"
    wb.save(out)
    return out
# === end generator ===
